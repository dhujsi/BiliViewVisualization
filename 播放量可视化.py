import re,os,openpyxl,numpy
from pylab import *
from openpyxl.chart import (
    LineChart,
    Reference,
    Series
)
from openpyxl.chart.axis import DateAxis
li=input("输入<ul>:")
if os.path.exists('D:/播放量可视化.xlsx')==True :
    wb=openpyxl.load_workbook('D:/播放量可视化.xlsx')
else:
    wb=openpyxl.Workbook()
ws=wb.create_sheet(0)
p1=re.compile(r'\d+\.\d+')
p2=re.compile(r'\d{4}-\d{2}-\d{2}')
m=p1.findall(li)
m=[float(i) for i in m]#激励收入
t=p2.findall(li)#日期
p=[i*272 for i in m]#换算得到的播放量，272播放/元是我根据自己视频数据算出的参数。
p2=list(numpy.log10(p))
t.insert(0,"合计")
t.append(str("date"))
m.insert(0,sum(m))
m.append(str("gain"))
p.insert(0,sum(p))
p.append(str("view"))
p2.insert(0,"？")
p2.append(str("log10(view)"))
final=list(zip(*[t,m,p,p2]))[::-1]
for i in range(len(final)):
    ws.append(final[i])
c1 = LineChart()
c1.title = "日期-播放量"
c1.style = 2
c1.x_axis.title = '日期'
c2 = LineChart()
c2.title = "日期-log10(播放量)"
c2.style = 2
c2.x_axis.title = '日期'
dates=Reference(
    ws,
    min_col=1,
    min_row=2,
    max_col=1,
    max_row=ws.max_row-1
)

data=Reference(
    ws,
    min_col=3,
    min_row=2,
    max_col=3,
    max_row=ws.max_row-1
)
lgdata=Reference(
    ws,
    min_col=4,
    min_row=2,
    max_col=4,
    max_row=ws.max_row-1
)
seriesObj1 = Series(data, title='播放量')
seriesObj2 = Series(lgdata, title='log10(播放量)')
c1.append(seriesObj1)
c1.set_categories(dates)
c2.append(seriesObj2)
c2.set_categories(dates)
ws.add_chart(c1, "F1")
ws.add_chart(c2, "F16")
wb.save('D:/播放量可视化.xlsx')
